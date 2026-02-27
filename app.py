import streamlit as st
import time
import re
import io
from zhipuai import ZhipuAI
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText

# --- 1. 安全配置：读取 API Key ---
try:
    API_KEY = st.secrets["ZHIPU_API_KEY"]
    client = ZhipuAI(api_key=API_KEY)
except Exception:
    st.error("❌ 未检测到 API Key。请在 .streamlit/secrets.toml 或云端 Secrets 中配置 ZHIPU_API_KEY。")
    st.stop()

# --- 2. 界面配置 ---
st.set_page_config(page_title="Marshall's AI Translator", page_icon="🌍", layout="wide")

st.title("🌍 Marshall's Excel workstation")
st.markdown("针对 **合并单元格**、**多Sheet** 进行了专项优化。")

# --- 3. 语言映射配置 ---
LANG_OPTIONS = {
    "简体中文": "Chinese",
    "英语": "English",
    "阿拉伯语": "Arabic",
    "法语": "French",
    "西班牙语": "Spanish",
    "德语": "German"
}

# 侧边栏设置
with st.sidebar:
    st.header("⚙️ 翻译配置")
    source_lang = st.selectbox("1. 原始语言 (源)", list(LANG_OPTIONS.keys()), index=0)
    target_lang = st.selectbox("2. 目标语言 (译)", list(LANG_OPTIONS.keys()), index=1) # 默认选阿拉伯语
    
    st.write("---")
    st.info(f"模式：从 **{source_lang}** 翻译至 **{target_lang}**")
    
    # 自动识别是否需要 RTL 布局
    is_rtl = st.checkbox("强制开启右至左 (RTL) 布局", value=True if "阿拉伯" in target_lang else False)

# --- 4. 核心引擎 ---
def translate_engine(text, src, tgt):
    if not text or str(text).strip() == "":
        return text
    # 过滤纯数字、物流单号、纯缩写
    if re.fullmatch(r'^[A-Z0-9\s\-_./()]+$', str(text).strip()):
        return text
    
    time.sleep(0.4) # 防频率报错
    try:
        response = client.chat.completions.create(
            model="glm-4",
            messages=[
                {"role": "system", "content": f"""
                    你是一个精通{src}和{tgt}的物流与IT专家。
                    任务：将内容翻译为{tgt}。
                    要求：保持术语(PUDO, UPS, Dangerous Goods, Maotai)和编号不变。
                    中英混装内容需合并翻译。只返回译文结果。
                """},
                {"role": "user", "content": str(text)}
            ],
            top_p=0.7, temperature=0.1,
        )
        return response.choices[0].message.content.strip()
    except:
        return text

# --- 5. 主逻辑区 ---
uploaded_file = st.file_uploader("上传您的 Excel 文件 (.xlsx)", type=["xlsx"])

if uploaded_file:
    if source_lang == target_lang:
        st.warning("⚠️ 原始语言和目标语言相同，请重新选择。")
    elif st.button("🚀 开始自动化翻译"):
        # 读取文件
        wb = load_workbook(filename=io.BytesIO(uploaded_file.read()))
        status_msg = st.empty()
        progress_bar = st.progress(0)
        
        sheet_names = wb.sheetnames
        for idx, name in enumerate(sheet_names):
            ws = wb[name]
            status_msg.info(f"正在翻译工作表: **{name}** ({idx+1}/{len(sheet_names)})")
            
            # 布局调整
            if is_rtl:
                ws.sheet_view.rightToLeft = True
            
            # 单元格遍历
            for row in ws.iter_rows():
                for cell in row:
                    # 跳过公式
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        continue
                    
                    # 翻译逻辑（含富文本）
                    if isinstance(cell.value, CellRichText):
                        new_rt = CellRichText()
                        for seg in cell.value:
                            if isinstance(seg, str):
                                new_rt.append(translate_engine(seg, source_lang, target_lang))
                            else:
                                seg.text = translate_engine(seg.text, source_lang, target_lang)
                                new_rt.append(seg)
                        cell.value = new_rt
                    elif isinstance(cell.value, str):
                        cell.value = translate_engine(cell.value, source_lang, target_lang)
            
            progress_bar.progress((idx + 1) / len(sheet_names))

        # 下载区域
        out_buffer = io.BytesIO()
        wb.save(out_buffer)
        st.success("🎉 所有 Sheet 翻译已完成！")
        st.download_button(
            label="💾 下载已翻译的文件",
            data=out_buffer.getvalue(),
            file_name=f"{LANG_OPTIONS[target_lang]}_{uploaded_file.name}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        )
